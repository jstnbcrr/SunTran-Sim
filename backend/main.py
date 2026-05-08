"""FastAPI backend for the SunTran Transit Simulation Tool."""

import io
import os
from typing import Optional

import pandas as pd
from fastapi import FastAPI, HTTPException, UploadFile, File, Depends, APIRouter
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from auth import (
    authenticate_user, create_access_token, get_current_user, get_current_admin,
    create_user, delete_user, list_usernames,
)
from route_loader import (
    load_stops,
    load_routes,
    load_ridership,
    load_employment_hubs,
    routes_to_dict,
    stops_to_dict,
)
from simulation_engine import build_transit_graph, compare_networks
from metrics import generate_accessibility_report, export_metrics_csv

app = FastAPI(title="SunTran Simulation API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# All routes on this router require a valid JWT token
protected = APIRouter(dependencies=[Depends(get_current_user)])

import sys as _sys
if getattr(_sys, "frozen", False):
    # Running as a PyInstaller bundle — data lives next to the .exe
    DATA_DIR = os.path.join(os.path.dirname(_sys.executable), "data")
else:
    DATA_DIR = os.environ.get("SUNTRAN_DATA_DIR") or os.path.join(os.path.dirname(__file__), "..", "data")

BACKUP_DIR = os.path.join(DATA_DIR, "backups")
OTP_ARCHIVE_DIR = os.path.join(DATA_DIR, "otp_archive")
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(OTP_ARCHIVE_DIR, exist_ok=True)

REQUIRED_CSV_COLUMNS = {
    "stops": {"stop_id", "stop_name", "latitude", "longitude"},
    "routes": {"route_id", "route_name", "stop_ids"},
    "ridership": {"route_id", "stop_id", "hour", "hourly_boardings", "hourly_alightings"},
    "employment_hubs": {"hub_name", "latitude", "longitude"},
}

# ── Route name aliases ────────────────────────────────────────────────────────
# Maps vendor route name variants → canonical name used in all aggregations.
# Add entries here whenever the vendor uses a different name for the same route.
ROUTE_NAME_ALIASES: dict[str, str] = {
    "Route 1 Red Cliffs (A)":                    "Route 1 Red Cliffs",
    "Route 1 Red Cliffs (B)":                    "Route 1 Red Cliffs",
    "Route 3 West Side Connector (Inbound)":     "Route 3 West Side Connector",
    "Route 3 West Side Connector (Outbound)":    "Route 3 West Side Connector",
}

# ── In-memory state ────────────────────────────────────────────────────────────
_stops: pd.DataFrame = pd.DataFrame()
_routes: list[dict] = []
_ridership: pd.DataFrame = pd.DataFrame()
_employment_hubs: pd.DataFrame = pd.DataFrame()
_otp: pd.DataFrame = pd.DataFrame()
_boardings_stop:        pd.DataFrame = pd.DataFrame()
_boardings_route:       pd.DataFrame = pd.DataFrame()
_boardings_dow:         pd.DataFrame = pd.DataFrame()
_boardings_month:       pd.DataFrame = pd.DataFrame()
_boardings_route_dow:   pd.DataFrame = pd.DataFrame()
_boardings_route_month: pd.DataFrame = pd.DataFrame()
_boardings_route_stop:  pd.DataFrame = pd.DataFrame()
_boardings_hour:        pd.DataFrame = pd.DataFrame()
_boardings_route_hour:  pd.DataFrame = pd.DataFrame()
_boardings_stop_month:  pd.DataFrame = pd.DataFrame()
_boardings_dow_month:   pd.DataFrame = pd.DataFrame()


def _save_routes():
    """Persist in-memory routes back to routes.csv so edits survive restarts."""
    import csv as csv_mod
    path = os.path.join(DATA_DIR, "routes.csv")
    with open(path, "w", newline="") as f:
        writer = csv_mod.DictWriter(f, fieldnames=["route_id", "route_name", "color", "stop_ids"])
        writer.writeheader()
        for r in _routes:
            writer.writerow({
                "route_id": r["route_id"],
                "route_name": r["route_name"],
                "color": r.get("color", "#3388ff"),
                "stop_ids": "|".join(r["stop_ids"]),
            })


def _load_boardings():
    global _boardings_stop, _boardings_route, _boardings_dow, _boardings_month, \
           _boardings_route_dow, _boardings_route_month, _boardings_route_stop, \
           _boardings_hour, _boardings_route_hour, \
           _boardings_stop_month, _boardings_dow_month, _boardings_route_dow_month
    for attr, fname in [
        ("_boardings_stop",             "boardings_by_stop.csv"),
        ("_boardings_route",            "boardings_by_route.csv"),
        ("_boardings_dow",              "boardings_by_dow.csv"),
        ("_boardings_month",            "boardings_by_month.csv"),
        ("_boardings_route_dow",        "boardings_by_route_dow.csv"),
        ("_boardings_route_month",      "boardings_by_route_month.csv"),
        ("_boardings_route_stop",       "boardings_by_route_stop.csv"),
        ("_boardings_hour",             "boardings_by_hour.csv"),
        ("_boardings_route_hour",       "boardings_by_route_hour.csv"),
        ("_boardings_stop_month",       "boardings_by_stop_month.csv"),
        ("_boardings_dow_month",        "boardings_by_dow_month.csv"),
        ("_boardings_route_dow_month",  "boardings_by_route_dow_month.csv"),
    ]:
        path = os.path.join(DATA_DIR, fname)
        globals()[attr] = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()


def _load_otp():
    global _otp
    otp_path = os.path.join(DATA_DIR, "otp.csv")
    if os.path.exists(otp_path):
        _otp = pd.read_csv(otp_path)
    else:
        _otp = pd.DataFrame()


@app.on_event("startup")
def _startup():
    """Bootstrap admin users then load data."""
    import json, bcrypt
    users_file = os.path.join(os.path.dirname(__file__), "users.json")
    users = {}
    if os.path.exists(users_file):
        try:
            with open(users_file) as f:
                users = json.load(f)
        except Exception:
            pass
    changed = False
    for suffix in ["", "2", "3", "4", "5"]:
        username = os.environ.get(f"ADMIN_USER{suffix}")
        password = os.environ.get(f"ADMIN_PASSWORD{suffix}")
        if username and password and username not in users:
            users[username] = bcrypt.hashpw(password.encode(), bcrypt.gensalt(12)).decode()
            print(f"Created admin user '{username}'")
            changed = True
    if changed or not os.path.exists(users_file):
        with open(users_file, "w") as f:
            json.dump(users, f, indent=2)

    # If still no users (no env vars set, no users.json), auto-create one
    # and print the generated password so the operator can log in immediately.
    from auth import ensure_default_user
    ensure_default_user()

    _load_data()


def _load_data():
    global _stops, _routes, _ridership, _employment_hubs
    try:
        _stops = load_stops()
    except Exception as e:
        print(f"WARNING: could not load stops.csv: {e}")
        _stops = pd.DataFrame(columns=["stop_id", "stop_name", "latitude", "longitude"])
    try:
        _routes = routes_to_dict(load_routes())
    except Exception as e:
        print(f"WARNING: could not load routes.csv: {e}")
        _routes = []
    try:
        _ridership = load_ridership()
    except Exception as e:
        print(f"WARNING: could not load ridership.csv: {e}")
        _ridership = pd.DataFrame()
    try:
        _employment_hubs = load_employment_hubs()
    except Exception as e:
        print(f"WARNING: could not load employment_hubs.csv: {e}")
        _employment_hubs = pd.DataFrame()
    _load_otp()
    _load_boardings()


# ── Pydantic models ────────────────────────────────────────────────────────────

class StopModel(BaseModel):
    stop_id: str
    stop_name: str
    latitude: float
    longitude: float


class RouteModel(BaseModel):
    route_id: str
    route_name: str
    color: str = "#3388ff"
    stop_ids: list[str]


class SimulationParams(BaseModel):
    proposed_routes: list[RouteModel]
    proposed_stops: Optional[list[StopModel]] = None
    walking_radius_miles: float = 0.25
    max_travel_minutes: float = 30.0
    average_speed_mph: float = 15.0
    dwell_time_minutes: float = 0.5
    transfer_penalty_minutes: float = 5.0
    highway_speed_mph: float = 55.0
    highway_threshold_miles: float = 1.0


# ── Public endpoints (no auth required) ───────────────────────────────────────

@app.post("/api/auth/login")
def login(form: OAuth2PasswordRequestForm = Depends()):
    """Exchange username + password for a JWT access token."""
    username = authenticate_user(form.username, form.password)
    if not username:
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    token = create_access_token(username)
    return {"access_token": token, "token_type": "bearer", "username": username}


@app.get("/health")
def health():
    return {"status": "ok", "stops": len(_stops), "routes": len(_routes)}


# ── Protected endpoints ────────────────────────────────────────────────────────

@protected.get("/api/stops")
def get_stops():
    return stops_to_dict(_stops)


@protected.get("/api/routes")
def get_routes():
    return _routes


@protected.get("/api/employment-hubs")
def get_employment_hubs():
    return _employment_hubs.to_dict(orient="records")


@protected.get("/api/ridership")
def get_ridership():
    return _ridership.to_dict(orient="records")


@protected.get("/api/boardings/by-stop")
def get_boardings_by_stop():
    return [] if _boardings_stop.empty else _boardings_stop.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-route")
def get_boardings_by_route():
    # Derive from accumulated route_month data so historical months stack correctly
    if not _boardings_route_month.empty and "route" in _boardings_route_month.columns:
        df = _boardings_route_month.groupby("route", sort=False).agg(
            total_in=("total_in", "sum"),
            total_out=("total_out", "sum"),
            unique_days=("unique_days", "sum"),
        ).reset_index()
        df["avg_daily_in"]  = (df["total_in"]  / df["unique_days"].replace(0, 1)).round(1)
        df["avg_daily_out"] = (df["total_out"] / df["unique_days"].replace(0, 1)).round(1)
        return df.fillna("").to_dict(orient="records")
    return [] if _boardings_route.empty else _boardings_route.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-dow")
def get_boardings_by_dow():
    return [] if _boardings_dow.empty else _boardings_dow.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-month")
def get_boardings_by_month():
    return [] if _boardings_month.empty else _boardings_month.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-route-dow")
def get_boardings_by_route_dow():
    return [] if _boardings_route_dow.empty else _boardings_route_dow.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-route-dow-month")
def get_boardings_by_route_dow_month():
    return [] if _boardings_route_dow_month.empty else _boardings_route_dow_month.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-route-month")
def get_boardings_by_route_month():
    return [] if _boardings_route_month.empty else _boardings_route_month.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-route-stop")
def get_boardings_by_route_stop():
    return [] if _boardings_route_stop.empty else _boardings_route_stop.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-hour")
def get_boardings_by_hour():
    return [] if _boardings_hour.empty else _boardings_hour.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-route-hour")
def get_boardings_by_route_hour():
    return [] if _boardings_route_hour.empty else _boardings_route_hour.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-stop-month")
def get_boardings_by_stop_month():
    return [] if _boardings_stop_month.empty else _boardings_stop_month.fillna("").to_dict(orient="records")

@protected.get("/api/boardings/by-dow-month")
def get_boardings_by_dow_month():
    return [] if _boardings_dow_month.empty else _boardings_dow_month.fillna("").to_dict(orient="records")

# Stubs for endpoints referenced in client.js — data not yet collected
@protected.get("/api/boardings/by-dow-hour")
def get_boardings_by_dow_hour():
    return []

@protected.get("/api/boardings/by-route-dow-hour")
def get_boardings_by_route_dow_hour():
    return []

@protected.get("/api/boardings/by-route-stop-hour")
def get_boardings_by_route_stop_hour():
    return []


@protected.get("/api/otp")
def get_otp():
    if _otp.empty:
        return []
    return _otp.fillna("").to_dict(orient="records")


@protected.post("/api/reload")
def reload_data():
    """Re-read all CSV files from disk."""
    _load_data()
    return {"status": "reloaded"}


# ── Data management helpers ────────────────────────────────────────────────────

def _backup_file(file_type: str) -> Optional[str]:
    """Copy the current CSV to backups/ with a timestamp. Returns backup filename or None."""
    src = os.path.join(DATA_DIR, f"{file_type}.csv")
    if not os.path.exists(src):
        return None
    from datetime import datetime
    import shutil
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{file_type}_{ts}.csv"
    shutil.copy2(src, os.path.join(BACKUP_DIR, backup_name))
    return backup_name


def _list_backups(file_type: str) -> list[dict]:
    """Return backup metadata for a given file type, newest first."""
    import glob as glob_mod
    pattern = os.path.join(BACKUP_DIR, f"{file_type}_*.csv")
    files = sorted(glob_mod.glob(pattern), reverse=True)
    result = []
    for path in files:
        fname = os.path.basename(path)
        stat = os.stat(path)
        result.append({
            "filename": fname,
            "size_bytes": stat.st_size,
            "created_at": stat.st_mtime,
        })
    return result


# ── CSV Upload ─────────────────────────────────────────────────────────────────

@protected.post("/api/upload/{file_type}")
async def upload_csv(file_type: str, file: UploadFile = File(...)):
    allowed = {"stops", "routes", "ridership", "employment_hubs"}
    if file_type not in allowed:
        raise HTTPException(400, f"file_type must be one of {allowed}")

    content = await file.read()

    # Validate the CSV structure before writing to disk
    try:
        df = pd.read_csv(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Could not parse CSV: {e}")

    required = REQUIRED_CSV_COLUMNS[file_type]
    missing = required - set(df.columns)
    if missing:
        raise HTTPException(
            400,
            f"CSV is missing required columns: {sorted(missing)}. "
            f"Expected: {sorted(required)}",
        )

    # Back up the current file before overwriting
    backup_name = _backup_file(file_type)

    dest = os.path.join(DATA_DIR, f"{file_type}.csv")
    with open(dest, "wb") as f:
        f.write(content)

    _load_data()
    return {"status": "uploaded", "file_type": file_type, "rows": len(df), "backup": backup_name}


@protected.get("/api/data/{file_type}/download")
def download_current(file_type: str):
    """Download the current CSV for a file type."""
    allowed = {"stops", "routes", "ridership", "employment_hubs"}
    if file_type not in allowed:
        raise HTTPException(400, f"file_type must be one of {allowed}")
    path = os.path.join(DATA_DIR, f"{file_type}.csv")
    if not os.path.exists(path):
        raise HTTPException(404, f"No {file_type}.csv found")
    return StreamingResponse(
        open(path, "rb"),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={file_type}.csv"},
    )


@protected.get("/api/data/{file_type}/backups")
def list_backups(file_type: str):
    """List available backups for a file type."""
    allowed = {"stops", "routes", "ridership", "employment_hubs"}
    if file_type not in allowed:
        raise HTTPException(400, f"file_type must be one of {allowed}")
    return {"file_type": file_type, "backups": _list_backups(file_type)}


@protected.post("/api/data/{file_type}/restore/{filename}")
def restore_backup(file_type: str, filename: str):
    """Restore a backup file as the active CSV."""
    allowed = {"stops", "routes", "ridership", "employment_hubs"}
    if file_type not in allowed:
        raise HTTPException(400, f"file_type must be one of {allowed}")
    # Safety: only allow filenames that match the expected pattern
    if not filename.startswith(f"{file_type}_") or not filename.endswith(".csv"):
        raise HTTPException(400, "Invalid backup filename")
    import shutil
    backup_path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(backup_path):
        raise HTTPException(404, f"Backup {filename} not found")
    # Back up the current file before restoring
    _backup_file(file_type)
    dest = os.path.join(DATA_DIR, f"{file_type}.csv")
    shutil.copy2(backup_path, dest)
    _load_data()
    return {"status": "restored", "file_type": file_type, "from": filename}


@protected.delete("/api/data/{file_type}/backup/{filename}")
def delete_backup(file_type: str, filename: str):
    """Delete a specific backup file."""
    allowed = {"stops", "routes", "ridership", "employment_hubs"}
    if file_type not in allowed:
        raise HTTPException(400, f"file_type must be one of {allowed}")
    if not filename.startswith(f"{file_type}_") or not filename.endswith(".csv"):
        raise HTTPException(400, "Invalid backup filename")
    backup_path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(backup_path):
        raise HTTPException(404, f"Backup {filename} not found")
    os.remove(backup_path)
    return {"status": "deleted", "filename": filename}


@protected.get("/api/data/{file_type}/backup/{filename}/download")
def download_backup(file_type: str, filename: str):
    """Download a specific backup file."""
    allowed = {"stops", "routes", "ridership", "employment_hubs"}
    if file_type not in allowed:
        raise HTTPException(400, f"file_type must be one of {allowed}")
    if not filename.startswith(f"{file_type}_") or not filename.endswith(".csv"):
        raise HTTPException(400, "Invalid backup filename")
    backup_path = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(backup_path):
        raise HTTPException(404, f"Backup {filename} not found")
    return StreamingResponse(
        open(backup_path, "rb"),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Boardings / OTP merge config ───────────────────────────────────────────────

BOARDINGS_FILES: dict[str, dict] = {
    "boardings_by_month": {
        "filename": "boardings_by_month.csv",
        "key_cols": ["month"],
        "date_col": "month",
    },
    "boardings_by_route": {
        "filename": "boardings_by_route.csv",
        "key_cols": ["route"],
        "date_col": None,
    },
    "boardings_by_route_stop": {
        "filename": "boardings_by_route_stop.csv",
        "key_cols": ["route", "address"],
        "date_col": None,
    },
    "boardings_by_stop": {
        "filename": "boardings_by_stop.csv",
        "key_cols": ["route", "stop_id"],
        "date_col": None,
    },
    "boardings_by_dow": {
        "filename": "boardings_by_dow.csv",
        "key_cols": ["day_num"],
        "date_col": None,
    },
    "boardings_by_route_dow": {
        "filename": "boardings_by_route_dow.csv",
        "key_cols": ["route", "day_num"],
        "date_col": None,
    },
    "boardings_by_route_month": {
        "filename": "boardings_by_route_month.csv",
        "key_cols": ["route", "month"],
        "date_col": "month",
    },
    "otp": {
        "filename": "otp.csv",
        "key_cols": ["route_name", "stop_name"],
        "date_col": None,
    },
    "boardings_by_hour": {
        "filename": "boardings_by_hour.csv",
        "key_cols": ["hour"],
        "date_col": None,
    },
    "boardings_by_route_hour": {
        "filename": "boardings_by_route_hour.csv",
        "key_cols": ["route", "hour"],
        "date_col": None,
    },
    "boardings_by_stop_month": {
        "filename": "boardings_by_stop_month.csv",
        "key_cols": ["route", "stop_id", "month"],
        "date_col": "month",
    },
    "boardings_by_dow_month": {
        "filename": "boardings_by_dow_month.csv",
        "key_cols": ["day_num", "month"],
        "date_col": "month",
    },
}


def _boardings_file_info(file_type: str) -> dict:
    """Return metadata (row count, date range, last modified) for a boardings file."""
    cfg = BOARDINGS_FILES.get(file_type, {})
    path = os.path.join(DATA_DIR, cfg.get("filename", ""))
    if not os.path.exists(path):
        return {"exists": False, "rows": 0, "date_range": None, "last_modified": None}
    df = pd.read_csv(path)
    info: dict = {
        "exists": True,
        "rows": len(df),
        "date_range": None,
        "last_modified": os.path.getmtime(path),
    }
    date_col = cfg.get("date_col")
    if date_col and date_col in df.columns:
        vals = df[date_col].dropna().astype(str)
        if not vals.empty:
            info["date_range"] = {"min": str(vals.min()), "max": str(vals.max())}
    return info


def _merge_boardings_df(
    existing: pd.DataFrame,
    incoming: pd.DataFrame,
    key_cols: list[str],
) -> tuple[pd.DataFrame, int, int]:
    """Upsert incoming into existing using key_cols.  Returns (merged, added, updated)."""
    if existing.empty:
        return incoming.copy(), len(incoming), 0

    valid_keys = [k for k in key_cols if k in existing.columns and k in incoming.columns]
    if not valid_keys:
        merged = pd.concat([existing, incoming], ignore_index=True)
        return merged, len(incoming), 0

    def make_key(df: pd.DataFrame) -> "pd.Series[str]":
        return df[valid_keys].astype(str).agg("|".join, axis=1)

    existing_key = make_key(existing)
    incoming_key = make_key(incoming)

    is_new = ~incoming_key.isin(existing_key)
    is_update = incoming_key.isin(existing_key)
    keep_existing = ~existing_key.isin(incoming_key[is_update])

    merged = pd.concat(
        [existing[keep_existing], incoming[is_update], incoming[is_new]],
        ignore_index=True,
    )
    return merged, int(is_new.sum()), int(is_update.sum())


def _additive_merge_df(
    existing: pd.DataFrame,
    incoming: pd.DataFrame,
    key_cols: list[str],
    sum_cols: list[str],
) -> tuple[pd.DataFrame, int, int]:
    """
    Like _merge_boardings_df but ADDS sum_cols for matching keys instead of replacing.
    Used for DOW and hourly aggregates that must accumulate across monthly imports.
    Returns (merged, added, updated).
    """
    if existing.empty:
        return incoming.copy(), len(incoming), 0

    valid_keys = [k for k in key_cols if k in existing.columns and k in incoming.columns]
    if not valid_keys:
        merged = pd.concat([existing, incoming], ignore_index=True)
        return merged, len(incoming), 0

    def make_key(df: pd.DataFrame) -> "pd.Series[str]":
        return df[valid_keys].astype(str).agg("|".join, axis=1)

    result = existing.copy()
    for col in sum_cols:
        if col not in result.columns:
            result[col] = 0

    existing_keys = make_key(result)
    incoming_keys = make_key(incoming)

    is_new    = ~incoming_keys.isin(existing_keys)
    is_update =  incoming_keys.isin(existing_keys)

    updated_count = 0
    for _, row in incoming[is_update].iterrows():
        k = "|".join(str(row[c]) for c in valid_keys)
        mask = existing_keys == k
        for col in sum_cols:
            if col in result.columns:
                result.loc[mask, col] = result.loc[mask, col].values + row.get(col, 0)
        updated_count += 1

    new_rows = incoming[is_new]
    if len(new_rows):
        result = pd.concat([result, new_rows], ignore_index=True)

    return result, len(new_rows), updated_count


def _canonical_route_id(name: str) -> str:
    """Convert any route name variant to a canonical R{N} ID (e.g. 'Route 3 ...' → 'R3')."""
    import re
    m = re.search(r"Route\s+(\d+)", str(name), re.IGNORECASE)
    return f"R{m.group(1)}" if m else str(name)


def _normalize_otp_df(df: pd.DataFrame) -> pd.DataFrame:
    """Apply route name aliases and canonical route_id to an OTP DataFrame in-place."""
    df = df.copy()
    if ROUTE_NAME_ALIASES and "route_name" in df.columns:
        df["route_name"] = df["route_name"].replace(ROUTE_NAME_ALIASES)
    if "route_name" in df.columns:
        df["route_id"] = df["route_name"].apply(_canonical_route_id)
    return df


def _parse_otp_excel(content: bytes) -> pd.DataFrame:
    """Parse OTP Excel (Route_OTP_By_Route format) into otp.csv column layout."""
    import re

    xls = pd.read_excel(io.BytesIO(content), sheet_name="COMBINED")
    xls.columns = [str(c).strip() for c in xls.columns]

    def direction(name: str) -> str:
        n = str(name)
        if "(A)" in n or "Outbound" in n:
            return "A"
        if "(B)" in n or "Inbound" in n:
            return "B"
        m = re.search(r"\(([^)]+)\)\s*$", n)
        return m.group(1) if m else ""

    col_map = {
        "Route Name": "route_name",
        "Stop Name": "stop_name",
        "Early%": "early_pct",
        "On-Time%": "ontime_pct",
        "Late%": "late_pct",
        "Average Schedule Deviation (minutes)": "avg_deviation",
        "Total Trips": "total_trips",
        "Order": "order",
    }
    available = {k: v for k, v in col_map.items() if k in xls.columns}
    df = xls.rename(columns=available)[list(available.values())].copy()
    # Apply alias normalization + canonical route_id before direction tagging
    df = _normalize_otp_df(df)
    df["direction"] = df["route_name"].apply(direction)

    out_cols = [
        "route_id", "route_name", "direction", "stop_name", "order",
        "early_pct", "ontime_pct", "late_pct", "avg_deviation", "total_trips",
    ]
    for c in out_cols:
        if c not in df.columns:
            df[c] = None
    return df[out_cols]


# ── Raw AVL/APC vendor file parsers ───────────────────────────────────────────

def _parse_avg_passenger_counts(content: bytes) -> dict:
    """
    Parse APC Average Passenger Counts CSV (vendor format with metadata header rows).
    Filters non-revenue routes (Route ID <= 0) and returns 6 aggregated DataFrames
    ready to merge into the internal boardings CSVs.
    """
    import csv as _csv

    text = content.decode("utf-8", errors="replace")
    lines = text.splitlines()

    # Find the data header row: it starts with 'Date,' and is followed by date-formatted rows
    data_start = None
    for i, line in enumerate(lines):
        if ("Date" in line and "Route ID" in line and "Stop ID" in line):
            data_start = i
            break
    if data_start is None:
        # Fallback: look for the line after "Average Passenger Counts" label
        for i, line in enumerate(lines):
            if "Average Passenger Counts" in line:
                for j in range(i + 1, min(i + 6, len(lines))):
                    if "Date" in lines[j] and "Route" in lines[j]:
                        data_start = j
                        break
                if data_start is not None:
                    break
    if data_start is None:
        raise ValueError(
            "Could not find the data header row. Expected a row containing "
            "'Date', 'Route ID', and 'Stop ID' after the report metadata."
        )

    # Collect data lines; stop at next blank-line-preceded section or end of file
    data_lines = [lines[data_start]]
    for line in lines[data_start + 1:]:
        stripped = line.strip()
        if not stripped:
            continue  # skip blank lines inside data block
        # Stop when we hit a non-data row (e.g. a quoted section label with no digits)
        first_char = stripped.lstrip('"')[:1]
        if stripped.startswith('"') and not first_char.isdigit() and "Route" not in stripped:
            break
        data_lines.append(line)

    df = pd.read_csv(io.StringIO("\n".join(data_lines)))
    df.columns = [str(c).strip().strip('"') for c in df.columns]

    # Filter non-revenue routes (Route ID <= 0)
    if "Route ID" in df.columns:
        df = df[pd.to_numeric(df["Route ID"], errors="coerce").fillna(-999) > 0].copy()
    if df.empty:
        raise ValueError("No revenue-route rows found (Route ID > 0 required).")

    # Rename vendor columns → internal names
    df = df.rename(columns={
        "Route Name":      "route",
        "Stop Name":       "address",
        "Stop ID":         "stop_id",
        "Total Count In":  "total_in",
        "Total Count Out": "total_out",
    })
    df["total_in"]  = pd.to_numeric(df["total_in"],  errors="coerce").fillna(0).astype(int)
    df["total_out"] = pd.to_numeric(df["total_out"], errors="coerce").fillna(0).astype(int)
    df["stop_id"]   = df["stop_id"].astype(str).str.strip()

    # Normalize route names using the alias table
    if ROUTE_NAME_ALIASES:
        df["route"] = df["route"].replace(ROUTE_NAME_ALIASES)

    # Parse date and derive temporal fields
    df["_date"] = pd.to_datetime(df["Date"], format="%m/%d/%Y", errors="coerce")
    df = df.dropna(subset=["_date"])
    df["month"]    = df["_date"].dt.strftime("%Y-%m")
    df["day_num"]  = df["_date"].dt.dayofweek  # 0 = Monday
    df["day_name"] = df["_date"].dt.day_name()

    results = {}

    # 1. boardings_by_route_month  (key: route, month) — accumulates across months
    rm = df.groupby(["route", "month"], sort=False).agg(
        total_in=("total_in", "sum"),
        total_out=("total_out", "sum"),
        unique_days=("_date", "nunique"),
    ).reset_index()
    rm["avg_daily_in"]  = (rm["total_in"]  / rm["unique_days"].replace(0, 1)).round(1)
    rm["avg_daily_out"] = (rm["total_out"] / rm["unique_days"].replace(0, 1)).round(1)
    results["boardings_by_route_month"] = rm

    # 2. boardings_by_month  (key: month)
    bm = df.groupby("month", sort=False).agg(
        total_in=("total_in", "sum"),
        total_out=("total_out", "sum"),
        unique_days=("_date", "nunique"),
    ).reset_index()
    bm["avg_daily_in"]  = (bm["total_in"]  / bm["unique_days"].replace(0, 1)).round(1)
    bm["avg_daily_out"] = (bm["total_out"] / bm["unique_days"].replace(0, 1)).round(1)
    results["boardings_by_month"] = bm

    # 3. boardings_by_route_dow  (key: route, day_num) — additive merge (totals accumulate)
    rdow = df.groupby(["route", "day_num", "day_name"], sort=False).agg(
        total_in=("total_in", "sum"),
        total_out=("total_out", "sum"),
        _ndays=("_date", "nunique"),
    ).reset_index()
    rdow["avg_in"]  = (rdow["total_in"]  / rdow["_ndays"].replace(0, 1)).round(1)
    rdow["avg_out"] = (rdow["total_out"] / rdow["_ndays"].replace(0, 1)).round(1)
    results["boardings_by_route_dow"] = rdow[
        ["route", "day_num", "day_name", "avg_in", "avg_out", "total_in", "total_out"]
    ]

    # 4. boardings_by_dow  (key: day_num)
    dow = df.groupby(["day_num", "day_name"], sort=False).agg(
        total_in=("total_in", "sum"),
        total_out=("total_out", "sum"),
        _ndays=("_date", "nunique"),
    ).reset_index()
    dow["avg_in"]  = (dow["total_in"]  / dow["_ndays"].replace(0, 1)).round(1)
    dow["avg_out"] = (dow["total_out"] / dow["_ndays"].replace(0, 1)).round(1)
    results["boardings_by_dow"] = dow[
        ["day_num", "day_name", "avg_in", "avg_out", "total_in", "total_out"]
    ]

    # 5. boardings_by_route_stop  (key: route, address)
    rs = df.groupby(["route", "address"], sort=False).agg(
        total_in=("total_in", "sum"),
        total_out=("total_out", "sum"),
        days=("_date", "nunique"),
    ).reset_index()
    rs["avg_daily_in"]  = (rs["total_in"]  / rs["days"].replace(0, 1)).round(2)
    rs["avg_daily_out"] = (rs["total_out"] / rs["days"].replace(0, 1)).round(2)
    results["boardings_by_route_stop"] = rs

    # 6. boardings_by_stop  (key: route, stop_id)
    bs = df.groupby(["route", "stop_id", "address"], sort=False).agg(
        total_in=("total_in", "sum"),
        total_out=("total_out", "sum"),
        days=("_date", "nunique"),
    ).reset_index()
    bs["avg_daily_in"]  = (bs["total_in"]  / bs["days"].replace(0, 1)).round(2)
    bs["avg_daily_out"] = (bs["total_out"] / bs["days"].replace(0, 1)).round(2)
    results["boardings_by_stop"] = bs

    # 7. boardings_by_stop_month  (key: route, stop_id, month)
    bsm = df.groupby(["route", "stop_id", "address", "month"], sort=False).agg(
        total_in=("total_in", "sum"),
        total_out=("total_out", "sum"),
        days=("_date", "nunique"),
    ).reset_index()
    bsm["avg_daily_in"]  = (bsm["total_in"]  / bsm["days"].replace(0, 1)).round(2)
    bsm["avg_daily_out"] = (bsm["total_out"] / bsm["days"].replace(0, 1)).round(2)
    results["boardings_by_stop_month"] = bsm

    # 8. boardings_by_dow_month  (key: day_num, month)
    dowm = df.groupby(["day_num", "day_name", "month"], sort=False).agg(
        total_in=("total_in", "sum"),
        _ndays=("_date", "nunique"),
    ).reset_index()
    dowm["avg_in"] = (dowm["total_in"] / dowm["_ndays"].replace(0, 1)).round(1)
    results["boardings_by_dow_month"] = dowm[["day_num", "day_name", "month", "avg_in", "total_in"]]

    # 9. boardings_by_route_dow_month  (key: route, day_num, month)
    rdm = df.groupby(["route", "day_num", "day_name", "month"], sort=False).agg(
        total_in=("total_in", "sum"),
        _ndays=("_date", "nunique"),
    ).reset_index()
    rdm["avg_in"] = (rdm["total_in"] / rdm["_ndays"].replace(0, 1)).round(1)
    results["boardings_by_route_dow_month"] = rdm[["route", "day_num", "day_name", "month", "avg_in", "total_in"]]

    return results


def _parse_hourly_apc(content: bytes) -> dict:
    """
    Parse HourlyApcCounts CSV (vendor wide-format with 24 hour columns).
    Returns boardings_by_hour (system-wide totals) and boardings_by_route_hour.
    """
    import csv as _csv

    text = content.decode("utf-8", errors="replace")
    lines = text.splitlines()
    HOURS = [f"{h:02d}:00" for h in range(24)]

    sys_in_vals  = None
    sys_out_vals = None
    route_in_start = None

    for i, line in enumerate(lines):
        if '"Total Count In"' in line or line.strip().startswith('"Total Count In"'):
            try:
                row = list(_csv.reader([line]))[0]
                vals = []
                for v in row[1:25]:
                    try:
                        vals.append(int(float(v)))
                    except (ValueError, TypeError):
                        vals.append(0)
                if len(vals) == 24:
                    sys_in_vals = vals
            except Exception:
                pass
        elif '"Total Count Out"' in line or line.strip().startswith('"Total Count Out"'):
            try:
                row = list(_csv.reader([line]))[0]
                vals = []
                for v in row[1:25]:
                    try:
                        vals.append(int(float(v)))
                    except (ValueError, TypeError):
                        vals.append(0)
                if len(vals) == 24:
                    sys_out_vals = vals
            except Exception:
                pass
        elif "Hourly Passenger Count In by Route" in line:
            route_in_start = i + 1  # header row is next line

    # Build system-wide boardings_by_hour
    bh_rows = [
        {
            "hour": h,
            "total_in":  sys_in_vals[h]  if sys_in_vals  else 0,
            "total_out": sys_out_vals[h] if sys_out_vals else 0,
        }
        for h in range(24)
    ]
    boardings_by_hour = pd.DataFrame(bh_rows)

    # Parse per-route hourly section
    route_hour_rows = []
    if route_in_start is not None:
        # First line is the header
        header_line = lines[route_in_start]
        try:
            header = [h.strip().strip('"') for h in list(_csv.reader([header_line]))[0]]
            hour_indices = {int(h.split(":")[0]): idx for idx, h in enumerate(header) if h in HOURS}

            for raw_line in lines[route_in_start + 1:]:
                stripped = raw_line.strip()
                if not stripped:
                    continue
                if stripped.startswith('"Hourly') or stripped.startswith('"OTP'):
                    break
                try:
                    row = list(_csv.reader([raw_line]))[0]
                    try:
                        route_id_val = int(float(row[0]))
                    except (ValueError, TypeError, IndexError):
                        continue
                    if route_id_val <= 0:
                        continue
                    route_name = row[1].strip().strip('"') if len(row) > 1 else ""
                    for hour, col_idx in hour_indices.items():
                        if col_idx < len(row):
                            try:
                                val = int(float(row[col_idx]))
                            except (ValueError, TypeError):
                                val = 0
                            route_hour_rows.append({
                                "route": route_name,
                                "hour": hour,
                                "total_in": val,
                            })
                except Exception:
                    continue
        except Exception:
            pass

    boardings_by_route_hour = (
        pd.DataFrame(route_hour_rows)
        if route_hour_rows
        else pd.DataFrame(columns=["route", "hour", "total_in"])
    )

    return {
        "boardings_by_hour": boardings_by_hour,
        "boardings_by_route_hour": boardings_by_route_hour,
    }


def _parse_otp_csv(content: bytes) -> pd.DataFrame:
    """
    Parse TripOTPByRouteAndStop CSV (multi-section vendor format).
    Prefers the 'OTP by Date, Route, and Stop' section (most granular).
    Falls back to 'OTP by Date and Route' if stop-level data is absent.
    Returns a DataFrame matching otp.csv schema.
    """
    text = content.decode("utf-8", errors="replace")
    lines = text.splitlines()

    def _find_data_header(section_label: str) -> Optional[int]:
        """Return line index of the data-column header following section_label."""
        for i, line in enumerate(lines):
            if section_label in line:
                for j in range(i + 1, min(i + 6, len(lines))):
                    if "Route ID" in lines[j] and lines[j].strip():
                        return j
        return None

    def _read_section(header_idx: int) -> pd.DataFrame:
        if header_idx is None:
            return pd.DataFrame()
        section_lines = [lines[header_idx]]
        i = header_idx + 1
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
            # Stop at next section header (quoted label with no leading digit)
            if (line.startswith('"') and not line.lstrip('"')[:1].isdigit()
                    and "Route" not in line and "Stop" not in line):
                break
            section_lines.append(lines[i])
            i += 1
        try:
            df = pd.read_csv(io.StringIO("\n".join(section_lines)))
            df.columns = [str(c).strip().strip('"') for c in df.columns]
            return df
        except Exception:
            return pd.DataFrame()

    def _direction(name: str) -> str:
        n = str(name)
        if "(A)" in n or "Outbound" in n:
            return "A"
        if "(B)" in n or "Inbound" in n:
            return "B"
        return ""

    # ── Stop-level section (preferred) ────────────────────────────────────────
    stop_hdr = _find_data_header("OTP by Date, Route, and Stop")
    if stop_hdr is not None:
        df = _read_section(stop_hdr)
        if not df.empty and "Route ID" in df.columns:
            df = df[pd.to_numeric(df["Route ID"], errors="coerce").fillna(-1) > 0].copy()
            df = df.rename(columns={
                "Route ID":   "route_id",
                "Route Name": "route_name",
                "Stop ID":    "stop_id",
                "Stop Name":  "stop_name",
                "Average Schedule Deviation (minutes)": "avg_dev",
                "Early Trips":   "early_trips",
                "On-Time Trips": "ontime_trips",
                "Late Trips":    "late_trips",
            })
            for col in ["early_trips", "ontime_trips", "late_trips"]:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
            df["avg_dev"] = pd.to_numeric(df.get("avg_dev", 0), errors="coerce").fillna(0)

            # Normalize route names + re-derive canonical route_id BEFORE groupby
            df = _normalize_otp_df(df)

            agg = df.groupby(["route_id", "route_name", "stop_name"], sort=False).agg(
                early_trips=("early_trips", "sum"),
                ontime_trips=("ontime_trips", "sum"),
                late_trips=("late_trips", "sum"),
                avg_deviation=("avg_dev", "mean"),
            ).reset_index()
            agg["total_trips"] = agg["early_trips"] + agg["ontime_trips"] + agg["late_trips"]
            total = agg["total_trips"].replace(0, 1)
            agg["early_pct"]  = (agg["early_trips"]  / total * 100).round(2)
            agg["ontime_pct"] = (agg["ontime_trips"] / total * 100).round(2)
            agg["late_pct"]   = (agg["late_trips"]   / total * 100).round(2)
            agg["avg_deviation"] = agg["avg_deviation"].round(3)

            return pd.DataFrame({
                "route_id":      agg["route_id"],
                "route_name":    agg["route_name"],
                "direction":     agg["route_name"].apply(_direction),
                "stop_name":     agg["stop_name"],
                "order":         0,
                "early_pct":     agg["early_pct"],
                "ontime_pct":    agg["ontime_pct"],
                "late_pct":      agg["late_pct"],
                "avg_deviation": agg["avg_deviation"],
                "total_trips":   agg["total_trips"],
            })

    # ── Route-level fallback ───────────────────────────────────────────────────
    route_hdr = _find_data_header("OTP by Date and Route")
    if route_hdr is not None:
        df = _read_section(route_hdr)
        if not df.empty and "Route ID" in df.columns:
            df = df[pd.to_numeric(df["Route ID"], errors="coerce").fillna(-1) > 0].copy()
            df = df.rename(columns={
                "Route ID": "route_id", "Route Name": "route_name",
                "Early Trip Stops": "early_trips",
                "On-Time Trip Stops": "ontime_trips",
                "Late Trip Stops": "late_trips",
            })
            for col in ["early_trips", "ontime_trips", "late_trips"]:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

            # Normalize route names + re-derive canonical route_id BEFORE groupby
            df = _normalize_otp_df(df)

            agg = df.groupby(["route_id", "route_name"], sort=False).agg(
                early_trips=("early_trips", "sum"),
                ontime_trips=("ontime_trips", "sum"),
                late_trips=("late_trips", "sum"),
            ).reset_index()
            agg["total_trips"] = agg["early_trips"] + agg["ontime_trips"] + agg["late_trips"]
            total = agg["total_trips"].replace(0, 1)
            agg["early_pct"]  = (agg["early_trips"]  / total * 100).round(2)
            agg["ontime_pct"] = (agg["ontime_trips"] / total * 100).round(2)
            agg["late_pct"]   = (agg["late_trips"]   / total * 100).round(2)

            return pd.DataFrame({
                "route_id":      agg["route_id"],
                "route_name":    agg["route_name"],
                "direction":     agg["route_name"].apply(_direction),
                "stop_name":     "",
                "order":         0,
                "early_pct":     agg["early_pct"],
                "ontime_pct":    agg["ontime_pct"],
                "late_pct":      agg["late_pct"],
                "avg_deviation": 0.0,
                "total_trips":   agg["total_trips"],
            })

    raise ValueError(
        "No OTP data section found. Expected 'OTP by Date, Route, and Stop' "
        "or 'OTP by Date and Route' section in the file."
    )


# ── Smart auto-detection ───────────────────────────────────────────────────────

# Detection rules: ordered most-specific → least-specific.
# Each entry: (file_type, required_cols_set, display_label)
_DETECT_RULES: list[tuple[str, set[str], str]] = [
    ("boardings_by_route_month", {"route", "month", "total_in"},             "Boardings by Route × Month"),
    ("boardings_by_route_stop",  {"route", "address", "total_in"},           "Boardings by Route × Stop"),
    ("boardings_by_route_dow",   {"route", "day_num", "day_name"},           "Boardings by Route × Day of Week"),
    ("boardings_by_month",       {"month", "total_in", "unique_days"},       "Boardings by Month"),
    ("boardings_by_stop",        {"stop_id", "address", "total_in"},         "Boardings by Stop"),
    ("boardings_by_dow",         {"day_num", "day_name", "total_in"},        "Boardings by Day of Week"),
    ("boardings_by_route",       {"route", "total_in", "unique_days"},       "Boardings by Route"),
    ("ridership",                {"hourly_boardings", "hourly_alightings"},  "Hourly Ridership"),
    ("stops",                    {"stop_id", "stop_name", "latitude", "longitude"}, "Bus Stops"),
    ("routes",                   {"route_id", "route_name", "stop_ids"},     "Routes"),
    ("employment_hubs",          {"hub_name", "estimated_workers"},          "Employment Hubs"),
    ("otp",                      {"early_pct", "ontime_pct", "late_pct"},    "On-Time Performance"),
]

_NETWORK_TYPES = {"stops", "routes", "ridership", "employment_hubs"}


def _detect_csv_type(cols: set[str]) -> tuple[Optional[str], str]:
    """Return (file_type, display_label) or (None, 'Unknown') for a CSV column set."""
    for file_type, required, label in _DETECT_RULES:
        if required.issubset(cols):
            return file_type, label
    return None, "Unknown — could not identify"


def _smart_import_file(content: bytes, filename: str, mode: str = "merge") -> dict:
    """
    Auto-detect and import a single file (CSV or XLSX).
    Returns a result dict describing what happened.
    """
    lower = filename.lower()

    # ── Excel → try OTP parse ──────────────────────────────────────────────────
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        try:
            incoming = _parse_otp_excel(content)
            file_type = "otp"
            label = "On-Time Performance (Excel)"
        except Exception as e:
            return {"filename": filename, "status": "error", "error": str(e),
                    "detected_as": "Excel", "label": "Excel"}

        period = _extract_period_from_filename(filename)
        _archive_otp(incoming, period)

        path = os.path.join(DATA_DIR, "otp.csv")
        backup_name = _backup_file("otp") if os.path.exists(path) else None
        if mode == "replace":
            final_df, added, updated = incoming, len(incoming), 0
        else:
            existing = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
            cfg = BOARDINGS_FILES["otp"]
            final_df, added, updated = _merge_boardings_df(existing, incoming, cfg["key_cols"])
        final_df.to_csv(path, index=False)
        _load_otp()
        return {"filename": filename, "status": "imported", "detected_as": file_type,
                "label": label, "rows_added": added, "rows_updated": updated,
                "total_rows": len(final_df), "backup": backup_name, "period": period}

    # ── CSV ────────────────────────────────────────────────────────────────────
    try:
        df = pd.read_csv(io.BytesIO(content))
    except Exception as e:
        return {"filename": filename, "status": "error", "error": f"Could not parse CSV: {e}",
                "detected_as": None, "label": "Unknown"}

    cols = set(df.columns)
    file_type, label = _detect_csv_type(cols)

    if file_type is None:
        return {"filename": filename, "status": "skipped",
                "detected_as": None, "label": "Unknown — could not identify",
                "rows": len(df), "columns": sorted(cols)}

    # Network core files (stops, routes, ridership, employment_hubs)
    if file_type in _NETWORK_TYPES:
        backup_name = _backup_file(file_type)
        dest = os.path.join(DATA_DIR, f"{file_type}.csv")
        with open(dest, "wb") as f:
            f.write(content)
        _load_data()
        return {"filename": filename, "status": "imported", "detected_as": file_type,
                "label": label, "rows_added": len(df), "rows_updated": 0,
                "total_rows": len(df), "backup": backup_name}

    # Boardings / OTP CSV files (merge mode)
    cfg = BOARDINGS_FILES.get(file_type, {})
    path = os.path.join(DATA_DIR, cfg.get("filename", f"{file_type}.csv"))
    backup_name = _backup_file(file_type) if os.path.exists(path) else None
    if mode == "replace":
        final_df, added, updated = df, len(df), 0
    else:
        existing = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
        final_df, added, updated = _merge_boardings_df(existing, df, cfg.get("key_cols", []))
    final_df.to_csv(path, index=False)
    _load_boardings()
    return {"filename": filename, "status": "imported", "detected_as": file_type,
            "label": label, "rows_added": added, "rows_updated": updated,
            "total_rows": len(final_df), "backup": backup_name}


@protected.post("/api/upload/smart/preview")
async def smart_import_preview(files: list[UploadFile] = File(...)):
    """Identify each file without writing anything to disk."""
    results = []
    for upload in files:
        content = await upload.read()
        lower = upload.filename.lower()
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            try:
                df = _parse_otp_excel(content)
                results.append({
                    "filename": upload.filename,
                    "detected_as": "otp",
                    "label": "On-Time Performance (Excel)",
                    "incoming_rows": len(df),
                    "status": "ready",
                })
            except Exception as e:
                results.append({"filename": upload.filename, "detected_as": None,
                                 "label": "Excel parse error", "status": "error", "error": str(e)})
        else:
            try:
                df = pd.read_csv(io.BytesIO(content))
            except Exception as e:
                results.append({"filename": upload.filename, "detected_as": None,
                                 "label": "CSV parse error", "status": "error", "error": str(e)})
                continue
            cols = set(df.columns)
            file_type, label = _detect_csv_type(cols)
            entry: dict = {
                "filename": upload.filename,
                "detected_as": file_type,
                "label": label,
                "incoming_rows": len(df),
                "status": "ready" if file_type else "unknown",
            }
            # For boardings files, add merge preview numbers
            if file_type and file_type not in _NETWORK_TYPES and file_type in BOARDINGS_FILES:
                cfg = BOARDINGS_FILES[file_type]
                path = os.path.join(DATA_DIR, cfg["filename"])
                existing = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
                _, added, updated = _merge_boardings_df(existing, df, cfg["key_cols"])
                entry["rows_to_add"] = added
                entry["rows_to_update"] = updated
                entry["existing_rows"] = len(existing)
                if cfg["date_col"] and cfg["date_col"] in df.columns:
                    vals = df[cfg["date_col"]].dropna().astype(str)
                    if not vals.empty:
                        entry["date_range"] = {"min": str(vals.min()), "max": str(vals.max())}
            results.append(entry)
    return {"files": results}


@protected.post("/api/upload/smart")
async def smart_import(files: list[UploadFile] = File(...), mode: str = "merge"):
    """Auto-detect and import all uploaded files in one request."""
    results = []
    for upload in files:
        content = await upload.read()
        result = _smart_import_file(content, upload.filename, mode)
        results.append(result)
    imported = sum(1 for r in results if r["status"] == "imported")
    skipped  = sum(1 for r in results if r["status"] == "skipped")
    errors   = sum(1 for r in results if r["status"] == "error")
    return {"status": "done", "imported": imported, "skipped": skipped,
            "errors": errors, "files": results}


# ── Boardings file info ─────────────────────────────────────────────────────────

@protected.get("/api/data/boardings/{file_type}/info")
def get_boardings_file_info(file_type: str):
    if file_type not in BOARDINGS_FILES:
        raise HTTPException(400, f"Unknown boardings file type: {file_type}")
    return _boardings_file_info(file_type)


@protected.get("/api/data/boardings/{file_type}/download")
def download_boardings_file(file_type: str):
    """Download the current CSV for a boardings/otp file type."""
    if file_type not in BOARDINGS_FILES:
        raise HTTPException(400, f"Unknown boardings file type: {file_type}")
    cfg = BOARDINGS_FILES[file_type]
    path = os.path.join(DATA_DIR, cfg["filename"])
    if not os.path.exists(path):
        raise HTTPException(404, f"No {cfg['filename']} found")
    return StreamingResponse(
        open(path, "rb"),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={cfg['filename']}"},
    )


# ── Month management ────────────────────────────────────────────────────────────

_MONTH_FILES = [
    "boardings_by_route_month.csv",
    "boardings_by_month.csv",
    "boardings_by_stop_month.csv",
    "boardings_by_dow_month.csv",
    "boardings_by_route_dow_month.csv",
]


@protected.get("/api/boardings/months")
def list_boardings_months():
    """Return all months present in boardings_by_route_month.csv."""
    path = os.path.join(DATA_DIR, "boardings_by_route_month.csv")
    if not os.path.exists(path):
        return {"months": []}
    try:
        df = pd.read_csv(path)
        if "month" not in df.columns:
            return {"months": []}
        months = sorted(df["month"].dropna().astype(str).unique().tolist())
        return {"months": months}
    except Exception:
        return {"months": []}


@protected.delete("/api/boardings/month/{month}")
def delete_boardings_month(month: str):
    """
    Remove all data for a given month (YYYY-MM) from all boardings files
    that contain a 'month' column.  Also recomputes boardings_by_route.csv.
    """
    import re as _re
    if not _re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(400, f"Invalid month format '{month}'. Expected YYYY-MM.")

    removed = {}
    for fname in _MONTH_FILES:
        path = os.path.join(DATA_DIR, fname)
        if not os.path.exists(path):
            continue
        try:
            df = pd.read_csv(path)
            if "month" not in df.columns:
                continue
            before = len(df)
            df = df[df["month"].astype(str) != month]
            after = len(df)
            if before != after:
                _backup_file(fname.replace(".csv", ""))
                df.to_csv(path, index=False)
                removed[fname] = before - after
        except Exception as e:
            print(f"WARNING: could not process {fname} for month deletion: {e}")

    # Recompute boardings_by_route.csv from remaining route_month data
    rm_path = os.path.join(DATA_DIR, "boardings_by_route_month.csv")
    route_path = os.path.join(DATA_DIR, "boardings_by_route.csv")
    if os.path.exists(rm_path):
        try:
            rm_df = pd.read_csv(rm_path)
            if not rm_df.empty and "route" in rm_df.columns:
                by_route = rm_df.groupby("route", sort=False).agg(
                    total_in=("total_in", "sum"),
                    total_out=("total_out", "sum"),
                    unique_days=("unique_days", "sum"),
                ).reset_index()
                by_route["avg_daily_in"]  = (by_route["total_in"]  / by_route["unique_days"].replace(0, 1)).round(1)
                by_route["avg_daily_out"] = (by_route["total_out"] / by_route["unique_days"].replace(0, 1)).round(1)
                by_route.to_csv(route_path, index=False)
        except Exception as e:
            print(f"WARNING: could not recompute boardings_by_route: {e}")

    _load_boardings()

    if not removed:
        raise HTTPException(404, f"Month '{month}' not found in any boardings file.")

    return {"status": "deleted", "month": month, "rows_removed": removed}


# ── Boardings merge preview ─────────────────────────────────────────────────────

@protected.post("/api/upload/boardings/{file_type}/preview")
async def preview_boardings_upload(file_type: str, file: UploadFile = File(...)):
    """Dry-run: return change summary without writing to disk."""
    if file_type not in BOARDINGS_FILES:
        raise HTTPException(400, f"Unknown boardings file type: {file_type}")
    content = await file.read()
    try:
        incoming = pd.read_csv(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Could not parse CSV: {e}")

    cfg = BOARDINGS_FILES[file_type]
    path = os.path.join(DATA_DIR, cfg["filename"])
    existing = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
    _, added, updated = _merge_boardings_df(existing, incoming, cfg["key_cols"])

    new_date_range = None
    if cfg["date_col"] and cfg["date_col"] in incoming.columns:
        vals = incoming[cfg["date_col"]].dropna().astype(str)
        if not vals.empty:
            new_date_range = {"min": str(vals.min()), "max": str(vals.max())}

    return {
        "file_type": file_type,
        "existing_rows": len(existing),
        "incoming_rows": len(incoming),
        "rows_to_add": added,
        "rows_to_update": updated,
        "new_date_range": new_date_range,
    }


# ── Boardings merge upload ──────────────────────────────────────────────────────

@protected.post("/api/upload/boardings/{file_type}")
async def upload_boardings(
    file_type: str, file: UploadFile = File(...), mode: str = "merge"
):
    """Upload boardings CSV.  mode=merge (upsert) or mode=replace (overwrite)."""
    if file_type not in BOARDINGS_FILES:
        raise HTTPException(400, f"Unknown boardings file type: {file_type}")
    content = await file.read()
    try:
        incoming = pd.read_csv(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Could not parse CSV: {e}")

    cfg = BOARDINGS_FILES[file_type]
    path = os.path.join(DATA_DIR, cfg["filename"])
    backup_name = _backup_file(file_type) if os.path.exists(path) else None

    if mode == "replace":
        final_df, added, updated = incoming, len(incoming), 0
    else:
        existing = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
        final_df, added, updated = _merge_boardings_df(existing, incoming, cfg["key_cols"])

    final_df.to_csv(path, index=False)
    _load_boardings()

    return {
        "status": "uploaded",
        "file_type": file_type,
        "mode": mode,
        "rows_added": added,
        "rows_updated": updated,
        "total_rows": len(final_df),
        "backup": backup_name,
    }


# ── OTP period helpers ─────────────────────────────────────────────────────────

_MONTH_ABBR = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}

def _extract_period_from_filename(filename: str) -> str:
    """Best-effort: pull a YYYY-MM period out of a filename.
    Falls back to today's YYYY-MM if nothing is found."""
    import re
    from datetime import date
    name = filename.lower()
    # Try YYYY-MM or MM-YYYY or YYYY_MM
    m = re.search(r"(20\d{2})[-_]?(0[1-9]|1[0-2])", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"(0[1-9]|1[0-2])[-_]?(20\d{2})", name)
    if m:
        return f"{m.group(2)}-{m.group(1)}"
    # Try month name + year
    for abbr, num in _MONTH_ABBR.items():
        if abbr in name:
            yr = re.search(r"20\d{2}", name)
            year = yr.group(0) if yr else str(date.today().year)
            return f"{year}-{num}"
    return date.today().strftime("%Y-%m")


def _archive_otp(df: pd.DataFrame, period: str) -> None:
    """Save a copy of an OTP DataFrame to the archive folder."""
    path = os.path.join(OTP_ARCHIVE_DIR, f"otp_{period}.csv")
    df.to_csv(path, index=False)


def _list_otp_periods() -> list[dict]:
    """Return sorted list of archived OTP periods with metadata."""
    periods = []
    for fname in os.listdir(OTP_ARCHIVE_DIR):
        if not fname.startswith("otp_") or not fname.endswith(".csv"):
            continue
        period = fname[4:-4]  # strip "otp_" prefix and ".csv" suffix
        fpath = os.path.join(OTP_ARCHIVE_DIR, fname)
        try:
            rows = sum(1 for _ in open(fpath)) - 1
        except Exception:
            rows = 0
        periods.append({"period": period, "rows": rows})
    return sorted(periods, key=lambda x: x["period"])


# ── OTP Excel upload ────────────────────────────────────────────────────────────

@protected.post("/api/upload/otp-excel")
async def upload_otp_excel(file: UploadFile = File(...), mode: str = "merge", period: str = None):
    """Accept OTP .xlsx (Route_OTP_By_Route format) and merge into otp.csv."""
    content = await file.read()
    try:
        incoming = _parse_otp_excel(content)
    except Exception as e:
        raise HTTPException(400, f"Could not parse OTP Excel: {e}")
    if incoming.empty:
        raise HTTPException(400, "No data found in COMBINED sheet")

    period = period or _extract_period_from_filename(file.filename or "")
    _archive_otp(incoming, period)

    path = os.path.join(DATA_DIR, "otp.csv")
    backup_name = _backup_file("otp") if os.path.exists(path) else None

    if mode == "replace":
        final_df, added, updated = incoming, len(incoming), 0
    else:
        existing = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
        cfg = BOARDINGS_FILES["otp"]
        final_df, added, updated = _merge_boardings_df(existing, incoming, cfg["key_cols"])

    final_df.to_csv(path, index=False)
    _load_otp()

    return {
        "status": "uploaded",
        "period": period,
        "rows_added": added,
        "rows_updated": updated,
        "total_rows": len(final_df),
        "backup": backup_name,
    }


# ── OTP archive endpoints ──────────────────────────────────────────────────────

@protected.get("/api/otp/periods")
def get_otp_periods():
    """List all archived OTP periods."""
    return _list_otp_periods()


@protected.get("/api/otp/period/{period}")
def get_otp_period(period: str):
    """Return OTP data for a specific archived period (format: YYYY-MM)."""
    path = os.path.join(OTP_ARCHIVE_DIR, f"otp_{period}.csv")
    if not os.path.exists(path):
        raise HTTPException(404, f"No OTP data found for period {period}")
    df = pd.read_csv(path)
    return df.fillna("").to_dict(orient="records")


@protected.delete("/api/otp/period/{period}")
def delete_otp_period(period: str):
    """Remove an archived OTP period."""
    path = os.path.join(OTP_ARCHIVE_DIR, f"otp_{period}.csv")
    if not os.path.exists(path):
        raise HTTPException(404, f"No OTP data found for period {period}")
    os.remove(path)
    return {"status": "deleted", "period": period}


# ── Raw AVL/APC vendor import endpoints ───────────────────────────────────────

_RAW_TYPES = {
    "avg-passenger": {
        "label":      "Average Passenger Counts",
        "info_file":  "boardings_by_route_month.csv",
        "info_date":  "month",
    },
    "otp-trip": {
        "label":      "Trip OTP by Route and Stop",
        "info_file":  "otp.csv",
        "info_date":  None,
    },
    "hourly-apc": {
        "label":      "Hourly APC Counts",
        "info_file":  "boardings_by_hour.csv",
        "info_date":  None,
    },
    "arrivals": {
        "label":      "Route Stop Arrival Times",
        "info_file":  "raw_arrivals.csv",
        "info_date":  None,
    },
}


@protected.get("/api/data/raw/{raw_type}/info")
def raw_import_info(raw_type: str):
    """Return metadata about what raw data has been imported for a given vendor file type."""
    if raw_type not in _RAW_TYPES:
        raise HTTPException(400, f"Unknown raw type: {raw_type}")
    cfg = _RAW_TYPES[raw_type]
    path = os.path.join(DATA_DIR, cfg["info_file"])
    if not os.path.exists(path):
        return {"exists": False, "rows": 0}
    try:
        df = pd.read_csv(path)
        info: dict = {"exists": True, "rows": len(df)}
        date_col = cfg.get("info_date")
        if date_col and date_col in df.columns:
            vals = df[date_col].dropna().astype(str)
            if not vals.empty:
                info["date_range"] = {"min": str(vals.min()), "max": str(vals.max())}
        return info
    except Exception:
        return {"exists": False, "rows": 0}


@protected.post("/api/upload/raw/{raw_type}/preview")
async def raw_upload_preview(raw_type: str, file: UploadFile = File(...)):
    """Dry-run: parse the vendor file and return a summary without writing anything."""
    if raw_type not in _RAW_TYPES:
        raise HTTPException(400, f"Unknown raw type: {raw_type}")
    content = await file.read()

    try:
        if raw_type == "avg-passenger":
            results = _parse_avg_passenger_counts(content)
            rm = results["boardings_by_route_month"]
            months = sorted(rm["month"].unique()) if "month" in rm.columns else []
            total_boardings = int(rm["total_in"].sum()) if "total_in" in rm.columns else 0

            # ── Route name validation ──────────────────────────────────────────
            known_routes: set[str] = set()
            rm_path = os.path.join(DATA_DIR, "boardings_by_route_month.csv")
            if os.path.exists(rm_path):
                try:
                    existing_rm = pd.read_csv(rm_path)
                    if "route" in existing_rm.columns:
                        known_routes.update(existing_rm["route"].dropna().astype(str).unique())
                except Exception:
                    pass
            known_routes.update(ROUTE_NAME_ALIASES.values())
            if "route" in rm.columns and known_routes:
                incoming_names = set(rm["route"].dropna().astype(str).unique())
                unrecognized = sorted(incoming_names - known_routes)
            else:
                unrecognized = []

            response: dict = {
                "incoming_rows":  int(sum(len(df) for df in results.values())),
                "routes":         int(rm["route"].nunique()) if "route" in rm.columns else 0,
                "stops":          len(results.get("boardings_by_stop", pd.DataFrame())),
                "total_boardings": total_boardings,
                "date_range":     {"min": months[0], "max": months[-1]} if months else None,
                "slots_updated":  list(results.keys()),
            }
            if unrecognized:
                response["warnings"] = [
                    {
                        "type": "unknown_route",
                        "message": f"Route name not recognized: '{name}'.",
                        "route_name": name,
                    }
                    for name in unrecognized
                ]
            return response

        elif raw_type == "otp-trip":
            df = _parse_otp_csv(content)
            routes = int(df["route_name"].nunique()) if "route_name" in df.columns else 0
            stops  = int(df["stop_name"].nunique())  if "stop_name"  in df.columns else 0

            # Validate OTP route names against canonical list
            known_routes_otp: set[str] = set()
            rm_path2 = os.path.join(DATA_DIR, "boardings_by_route_month.csv")
            if os.path.exists(rm_path2):
                try:
                    existing_rm2 = pd.read_csv(rm_path2)
                    if "route" in existing_rm2.columns:
                        known_routes_otp.update(existing_rm2["route"].dropna().astype(str).unique())
                except Exception:
                    pass
            known_routes_otp.update(ROUTE_NAME_ALIASES.values())

            otp_warnings = []
            if "route_name" in df.columns and known_routes_otp:
                incoming_names = set(df["route_name"].dropna().astype(str).unique())
                for name in sorted(incoming_names - known_routes_otp):
                    otp_warnings.append({
                        "type": "unknown_route",
                        "message": f"Route name not recognized: '{name}'.",
                        "route_name": name,
                    })

            otp_response = {
                "incoming_rows": len(df),
                "routes":        routes,
                "stops":         stops,
                "slots_updated": ["otp"],
            }
            if otp_warnings:
                otp_response["warnings"] = otp_warnings
            return otp_response
        elif raw_type == "hourly-apc":
            results = _parse_hourly_apc(content)
            bh = results["boardings_by_hour"]
            total = int(bh["total_in"].sum()) if not bh.empty else 0
            return {
                "incoming_rows":  len(bh),
                "total_boardings": total,
                "slots_updated":  ["boardings_by_hour", "boardings_by_route_hour"],
            }
        elif raw_type == "arrivals":
            # Just count rows (large file — don't fully parse)
            text  = content.decode("utf-8", errors="replace")
            nrows = max(0, text.count("\n") - 8)  # subtract ~8 header rows
            return {
                "incoming_rows": nrows,
                "slots_updated": ["raw_arrivals"],
            }
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")


@protected.post("/api/upload/raw/{raw_type}")
async def raw_upload(raw_type: str, file: UploadFile = File(...), period: str = None):
    """Parse a vendor-format file and merge it into the internal boardings/OTP datasets."""
    if raw_type not in _RAW_TYPES:
        raise HTTPException(400, f"Unknown raw type: {raw_type}")
    content = await file.read()
    backup_names = []

    try:
        if raw_type == "avg-passenger":
            results = _parse_avg_passenger_counts(content)

            # If any route aliases were applied, drop stale canonical records for
            # the affected (route, month) combinations before merging.
            if ROUTE_NAME_ALIASES:
                canonical_routes = set(ROUTE_NAME_ALIASES.values())
                incoming_rm = results.get("boardings_by_route_month", pd.DataFrame())
                if not incoming_rm.empty and "route" in incoming_rm.columns and "month" in incoming_rm.columns:
                    affected_months = incoming_rm[
                        incoming_rm["route"].isin(canonical_routes)
                    ]["month"].unique()
                    if len(affected_months):
                        for fname in ["boardings_by_route_month.csv", "boardings_by_stop_month.csv"]:
                            fpath = os.path.join(DATA_DIR, fname)
                            if not os.path.exists(fpath):
                                continue
                            edf = pd.read_csv(fpath)
                            if "route" in edf.columns and "month" in edf.columns:
                                drop_mask = (
                                    edf["route"].isin(canonical_routes) &
                                    edf["month"].isin(affected_months)
                                )
                                if drop_mask.any():
                                    edf = edf[~drop_mask]
                                    edf.to_csv(fpath, index=False)

            # Write each aggregated DataFrame, merging with existing data
            merge_cfg = {
                "boardings_by_route_month": (["route", "month"],            False),
                "boardings_by_month":        (["month"],                    False),
                "boardings_by_route_stop":   (["route", "address"],         False),
                "boardings_by_stop":         (["route", "stop_id"],         False),
                "boardings_by_stop_month":   (["route", "stop_id", "month"],False),
                "boardings_by_dow_month":         (["day_num", "month"],          False),
                "boardings_by_route_dow_month":   (["route", "day_num", "month"], False),
                # DOW files use additive merge so totals accumulate across months
                "boardings_by_route_dow":    (["route", "day_num"],         True),
                "boardings_by_dow":          (["day_num"],                  True),
            }
            totals = {"rows_added": 0, "rows_updated": 0}
            for name, (keys, additive) in merge_cfg.items():
                incoming_df = results.get(name)
                if incoming_df is None or incoming_df.empty:
                    continue
                path = os.path.join(DATA_DIR, f"{name}.csv")
                backup_names.append(_backup_file(name) if os.path.exists(path) else None)
                existing_df = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
                if additive:
                    sum_cols = [c for c in ["total_in", "total_out"] if c in incoming_df.columns]
                    final_df, added, updated = _additive_merge_df(
                        existing_df, incoming_df, keys, sum_cols
                    )
                    # Recompute averages after additive merge (DOW files)
                    if "total_in" in final_df.columns:
                        if "total_out" in final_df.columns:
                            # Recompute avg_in / avg_out from accumulated totals + count
                            # Use n_days if present, else infer from ratio
                            if "avg_in" in final_df.columns:
                                # Store totals as source of truth; avg is display only
                                ndays_inferred = (
                                    final_df["total_in"] / final_df["avg_in"].replace(0, 1)
                                ).round().clip(lower=1)
                                final_df["avg_in"]  = (final_df["total_in"]  / ndays_inferred).round(1)
                                final_df["avg_out"] = (final_df["total_out"] / ndays_inferred).round(1)
                else:
                    final_df, added, updated = _merge_boardings_df(existing_df, incoming_df, keys)
                final_df.to_csv(path, index=False)
                totals["rows_added"]   += added
                totals["rows_updated"] += updated

            # Recompute boardings_by_route.csv from accumulated route_month data
            rm_path = os.path.join(DATA_DIR, "boardings_by_route_month.csv")
            if os.path.exists(rm_path):
                rm_df = pd.read_csv(rm_path)
                if not rm_df.empty and "route" in rm_df.columns:
                    by_route = rm_df.groupby("route", sort=False).agg(
                        total_in=("total_in", "sum"),
                        total_out=("total_out", "sum"),
                        unique_days=("unique_days", "sum"),
                    ).reset_index()
                    by_route["avg_daily_in"]  = (by_route["total_in"]  / by_route["unique_days"].replace(0, 1)).round(1)
                    by_route["avg_daily_out"] = (by_route["total_out"] / by_route["unique_days"].replace(0, 1)).round(1)
                    by_route.to_csv(os.path.join(DATA_DIR, "boardings_by_route.csv"), index=False)

            _load_boardings()
            return {
                "status": "imported",
                "raw_type": raw_type,
                "rows_added": totals["rows_added"],
                "rows_updated": totals["rows_updated"],
                "backups": [b for b in backup_names if b],
            }

        elif raw_type == "otp-trip":
            incoming_df = _parse_otp_csv(content)
            if incoming_df.empty:
                raise HTTPException(400, "No OTP data parsed from file.")
            period = period or _extract_period_from_filename(file.filename or "")
            _archive_otp(incoming_df, period)
            path = os.path.join(DATA_DIR, "otp.csv")
            backup = _backup_file("otp") if os.path.exists(path) else None
            existing_df = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
            cfg = BOARDINGS_FILES["otp"]
            final_df, added, updated = _merge_boardings_df(existing_df, incoming_df, cfg["key_cols"])
            final_df.to_csv(path, index=False)
            _load_otp()
            return {
                "status": "imported",
                "raw_type": raw_type,
                "period": period,
                "rows_added": added,
                "rows_updated": updated,
                "total_rows": len(final_df),
                "backup": backup,
            }

        elif raw_type == "hourly-apc":
            results = _parse_hourly_apc(content)

            for name, keys in [
                ("boardings_by_hour",       ["hour"]),
                ("boardings_by_route_hour", ["route", "hour"]),
            ]:
                incoming_df = results.get(name)
                if incoming_df is None or incoming_df.empty:
                    continue
                path = os.path.join(DATA_DIR, f"{name}.csv")
                _backup_file(name) if os.path.exists(path) else None
                existing_df = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
                sum_cols = ["total_in", "total_out"] if "total_out" in incoming_df.columns else ["total_in"]
                final_df, _, _ = _additive_merge_df(existing_df, incoming_df, keys, sum_cols)
                final_df.to_csv(path, index=False)

            _load_boardings()
            return {"status": "imported", "raw_type": raw_type}

        elif raw_type == "arrivals":
            # Store raw file as-is (no transformation)
            path = os.path.join(DATA_DIR, "raw_arrivals.csv")
            with open(path, "wb") as f:
                f.write(content)
            return {"status": "imported", "raw_type": raw_type}

    except ValueError as e:
        raise HTTPException(400, str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Import failed: {e}")


# ── Structured import: template definitions ────────────────────────────────────
#
# Each slot defines the exact columns required, example values, a human label,
# a description, and how the data is stored / reloaded.

IMPORT_SLOTS: dict[str, dict] = {
    "boardings_by_month": {
        "label":   "Monthly Ridership Summary",
        "desc":    "Total boardings and alightings aggregated per calendar month.",
        "columns": ["month", "total_in", "total_out", "unique_days", "avg_daily_in", "avg_daily_out"],
        "example": ["2026-03", 31500, 32100, 22, 1431.8, 1459.1],
        "filename": "boardings_by_month.csv",
        "key_cols": ["month"],
        "date_col": "month",
        "reload": "_load_boardings",
        "group": "ridership",
    },
    "boardings_by_route_month": {
        "label":   "Route × Month Ridership",
        "desc":    "Per-route boardings broken down by month.",
        "columns": ["route", "month", "total_in", "total_out", "unique_days", "avg_daily_in", "avg_daily_out"],
        "example": ["Route 2 Riverside", "2026-03", 8200, 9100, 22, 372.7, 413.6],
        "filename": "boardings_by_route_month.csv",
        "key_cols": ["route", "month"],
        "date_col": "month",
        "reload": "_load_boardings",
        "group": "ridership",
    },
    "boardings_by_route": {
        "label":   "Route Totals",
        "desc":    "Cumulative boardings per route across all recorded days.",
        "columns": ["route", "total_in", "total_out", "unique_days", "avg_daily_in", "avg_daily_out"],
        "example": ["Route 2 Riverside", 80475, 93623, 336, 239.5, 278.6],
        "filename": "boardings_by_route.csv",
        "key_cols": ["route"],
        "date_col": None,
        "reload": "_load_boardings",
        "group": "ridership",
    },
    "boardings_by_route_stop": {
        "label":   "Route × Stop Boardings",
        "desc":    "Boardings and alightings per stop per route.",
        "columns": ["route", "address", "total_in", "total_out", "avg_daily_in", "avg_daily_out", "days"],
        "example": ["Route 2 Riverside", "Transit Center", 37021, 56856, 110.2, 169.2, 336],
        "filename": "boardings_by_route_stop.csv",
        "key_cols": ["route", "address"],
        "date_col": None,
        "reload": "_load_boardings",
        "group": "ridership",
    },
    "boardings_by_stop": {
        "label":   "Stop-Level Boardings",
        "desc":    "Boardings at every individual stop across all routes.",
        "columns": ["route", "stop_id", "address", "total_in", "total_out", "avg_daily_in", "avg_daily_out", "days"],
        "example": ["Route 1 Red Cliffs (B)", "143486", "449 North 2450 East", 946, 1017, 2.88, 3.09, 329],
        "filename": "boardings_by_stop.csv",
        "key_cols": ["route", "stop_id"],
        "date_col": None,
        "reload": "_load_boardings",
        "group": "ridership",
    },
    "boardings_by_dow": {
        "label":   "Day-of-Week Ridership",
        "desc":    "Average and total boardings for each day of the week.",
        "columns": ["day_num", "day_name", "avg_in", "avg_out", "total_in", "total_out"],
        "example": [0, "Monday", 7.87, 8.56, 69691, 75748],
        "filename": "boardings_by_dow.csv",
        "key_cols": ["day_num"],
        "date_col": None,
        "reload": "_load_boardings",
        "group": "ridership",
    },
    "boardings_by_route_dow": {
        "label":   "Route × Day-of-Week",
        "desc":    "Per-route ridership broken down by day of the week.",
        "columns": ["route", "day_num", "day_name", "avg_in", "avg_out", "total_in", "total_out"],
        "example": ["Route 1 Red Cliffs (A)", 0, "Monday", 10.5, 10.8, 6259, 6435],
        "filename": "boardings_by_route_dow.csv",
        "key_cols": ["route", "day_num"],
        "date_col": None,
        "reload": "_load_boardings",
        "group": "ridership",
    },
    "otp": {
        "label":   "On-Time Performance",
        "desc":    "Early / on-time / late percentages per stop per route.",
        "columns": ["route_id", "route_name", "direction", "stop_name", "order",
                    "early_pct", "ontime_pct", "late_pct", "avg_deviation", "total_trips"],
        "example": ["R1", "Route 1 Red Cliffs (A)", "A", "Transit Center", 1,
                    45.98, 53.69, 0.33, 1.39, 2414],
        "filename": "otp.csv",
        "key_cols": ["route_name", "stop_name"],
        "date_col": None,
        "reload": "_load_otp",
        "group": "otp",
    },
    "stops": {
        "label":   "Bus Stops",
        "desc":    "All bus stop locations used in routing and the map.",
        "columns": ["stop_id", "stop_name", "latitude", "longitude"],
        "example": ["143486", "449 North 2450 East", 37.1139, -113.5401],
        "filename": "stops.csv",
        "key_cols": ["stop_id"],
        "date_col": None,
        "reload": "_load_data",
        "group": "network",
    },
    "routes": {
        "label":   "Routes",
        "desc":    "Route definitions with ordered stop lists.",
        "columns": ["route_id", "route_name", "color", "stop_ids"],
        "example": ["R1", "Route 1 Red Cliffs (A)", "#e74c3c", "143486|143487|143488"],
        "filename": "routes.csv",
        "key_cols": ["route_id"],
        "date_col": None,
        "reload": "_load_data",
        "group": "network",
    },
    "employment_hubs": {
        "label":   "Employment Hubs",
        "desc":    "Major employment destinations for accessibility analysis.",
        "columns": ["hub_name", "latitude", "longitude", "estimated_workers"],
        "example": ["Intermountain Health", 37.1052, -113.5841, 1200],
        "filename": "employment_hubs.csv",
        "key_cols": ["hub_name"],
        "date_col": None,
        "reload": "_load_data",
        "group": "network",
    },
}


def _generate_template(slot_key: str) -> bytes:
    """Return a styled .xlsx template file for the given import slot."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    slot = IMPORT_SLOTS[slot_key]
    wb = Workbook()
    ws = wb.active
    ws.title = slot_key

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    example_fill = PatternFill("solid", fgColor="1E2A3A")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    example_font = Font(color="A0B4C8", size=10, italic=True)
    thin = Border(
        left=Side(style="thin", color="2C4A6B"),
        right=Side(style="thin", color="2C4A6B"),
        top=Side(style="thin", color="2C4A6B"),
        bottom=Side(style="thin", color="2C4A6B"),
    )

    # Row 1: headers
    for col_idx, col_name in enumerate(slot["columns"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = max(14, len(col_name) + 4)

    # Row 2: example data
    for col_idx, value in enumerate(slot["example"], start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = example_font
        cell.fill = example_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/api/templates/{slot_key}")
def download_template(slot_key: str):
    """Return a blank .xlsx import template for the given slot."""
    if slot_key not in IMPORT_SLOTS:
        raise HTTPException(400, f"Unknown import slot: {slot_key}")
    xlsx_bytes = _generate_template(slot_key)
    slot = IMPORT_SLOTS[slot_key]
    fname = slot["filename"].replace(".csv", "_template.xlsx")
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@protected.get("/api/import/slots")
def list_import_slots():
    """Return slot metadata (without example data) for the frontend."""
    return {
        k: {
            "label":    v["label"],
            "desc":     v["desc"],
            "columns":  v["columns"],
            "filename": v["filename"],
            "date_col": v["date_col"],
            "group":    v["group"],
        }
        for k, v in IMPORT_SLOTS.items()
    }


@protected.get("/api/import/slots/{slot_key}/info")
def slot_current_info(slot_key: str):
    """Row count + date range for a slot's current data file."""
    if slot_key not in IMPORT_SLOTS:
        raise HTTPException(400, f"Unknown slot: {slot_key}")
    slot = IMPORT_SLOTS[slot_key]
    path = os.path.join(DATA_DIR, slot["filename"])
    if not os.path.exists(path):
        return {"exists": False, "rows": 0, "date_range": None}
    df = pd.read_csv(path)
    info: dict = {"exists": True, "rows": len(df), "date_range": None,
                  "last_modified": os.path.getmtime(path)}
    if slot["date_col"] and slot["date_col"] in df.columns:
        vals = df[slot["date_col"]].dropna().astype(str)
        if not vals.empty:
            info["date_range"] = {"min": str(vals.min()), "max": str(vals.max())}
    return info


@protected.post("/api/import/slots/{slot_key}/preview")
async def slot_upload_preview(slot_key: str, file: UploadFile = File(...)):
    """Validate columns and return merge diff without writing."""
    if slot_key not in IMPORT_SLOTS:
        raise HTTPException(400, f"Unknown slot: {slot_key}")
    slot = IMPORT_SLOTS[slot_key]
    content = await file.read()

    try:
        df = pd.read_excel(io.BytesIO(content)) if file.filename.endswith((".xlsx", ".xls")) \
             else pd.read_csv(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    # Strict column validation
    required = set(slot["columns"])
    actual   = set(df.columns)
    missing  = required - actual
    extra    = actual - required
    if missing:
        raise HTTPException(400, {
            "error": "column_mismatch",
            "message": f"File is missing required columns: {sorted(missing)}",
            "missing": sorted(missing),
            "extra":   sorted(extra),
            "expected": slot["columns"],
            "got": sorted(actual),
        })

    path = os.path.join(DATA_DIR, slot["filename"])
    existing = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
    _, added, updated = _merge_boardings_df(existing, df, slot["key_cols"])

    date_range = None
    if slot["date_col"] and slot["date_col"] in df.columns:
        vals = df[slot["date_col"]].dropna().astype(str)
        if not vals.empty:
            date_range = {"min": str(vals.min()), "max": str(vals.max())}

    return {
        "slot_key": slot_key,
        "existing_rows": len(existing),
        "incoming_rows": len(df),
        "rows_to_add":    added,
        "rows_to_update": updated,
        "date_range": date_range,
        "columns_ok": True,
    }


@protected.post("/api/import/slots/{slot_key}")
async def slot_upload(slot_key: str, file: UploadFile = File(...), mode: str = "merge"):
    """Validate and import a file into the given slot."""
    if slot_key not in IMPORT_SLOTS:
        raise HTTPException(400, f"Unknown slot: {slot_key}")
    slot = IMPORT_SLOTS[slot_key]
    content = await file.read()

    try:
        df = pd.read_excel(io.BytesIO(content)) if file.filename.endswith((".xlsx", ".xls")) \
             else pd.read_csv(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    required = set(slot["columns"])
    missing  = required - set(df.columns)
    if missing:
        raise HTTPException(400, {
            "error": "column_mismatch",
            "message": f"Missing columns: {sorted(missing)}",
            "missing": sorted(missing),
            "expected": slot["columns"],
        })

    # Keep only the expected columns, in order
    df = df[slot["columns"]]

    path = os.path.join(DATA_DIR, slot["filename"])
    backup_name = _backup_file(slot_key) if os.path.exists(path) else None

    if mode == "replace" or slot["group"] == "network":
        final_df, added, updated = df, len(df), 0
    else:
        existing = pd.read_csv(path) if os.path.exists(path) else pd.DataFrame()
        final_df, added, updated = _merge_boardings_df(existing, df, slot["key_cols"])

    final_df.to_csv(path, index=False)

    # Reload the right in-memory state
    reload_fn = slot["reload"]
    if reload_fn == "_load_data":
        _load_data()
    elif reload_fn == "_load_boardings":
        _load_boardings()
    elif reload_fn == "_load_otp":
        _load_otp()

    return {
        "status": "imported",
        "slot_key": slot_key,
        "mode": mode,
        "rows_added": added,
        "rows_updated": updated,
        "total_rows": len(final_df),
        "backup": backup_name,
    }


# ── Route CRUD ─────────────────────────────────────────────────────────────────

@protected.post("/api/routes")
def add_route(route: RouteModel):
    global _routes
    if any(r["route_id"] == route.route_id for r in _routes):
        raise HTTPException(409, f"Route {route.route_id} already exists")
    _routes.append(route.model_dump())
    _save_routes()
    return route


@protected.put("/api/routes/{route_id}")
def update_route(route_id: str, route: RouteModel):
    global _routes
    for i, r in enumerate(_routes):
        if r["route_id"] == route_id:
            _routes[i] = route.model_dump()
            _save_routes()
            return route
    raise HTTPException(404, f"Route {route_id} not found")


@protected.delete("/api/routes/{route_id}")
def delete_route(route_id: str):
    global _routes
    before = len(_routes)
    _routes = [r for r in _routes if r["route_id"] != route_id]
    if len(_routes) == before:
        raise HTTPException(404, f"Route {route_id} not found")
    _save_routes()
    return {"deleted": route_id}


# ── Metrics ────────────────────────────────────────────────────────────────────

@protected.get("/api/metrics")
def get_metrics():
    G = build_transit_graph(_routes, _stops)
    report = generate_accessibility_report(
        G, _stops, _employment_hubs, _routes, _ridership
    )
    return report


@protected.get("/api/metrics/export")
def export_metrics():
    G = build_transit_graph(_routes, _stops)
    report = generate_accessibility_report(
        G, _stops, _employment_hubs, _routes, _ridership
    )
    csv_content = export_metrics_csv(report)
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=transit_metrics.csv"},
    )


# ── Simulation ─────────────────────────────────────────────────────────────────

@protected.post("/api/simulate")
def run_simulation(params: SimulationParams):
    if params.proposed_stops:
        extra = pd.DataFrame([s.model_dump() for s in params.proposed_stops])
        merged_stops = pd.concat([_stops, extra], ignore_index=True).drop_duplicates(
            subset="stop_id"
        )
    else:
        merged_stops = _stops.copy()

    known_stop_ids = set(merged_stops["stop_id"].astype(str))

    # Validate proposed routes before building the graph
    for route in params.proposed_routes:
        if len(route.stop_ids) < 2:
            raise HTTPException(
                400, f"Route '{route.route_id}' must have at least 2 stops."
            )
        unknown = [sid for sid in route.stop_ids if sid not in known_stop_ids]
        if unknown:
            raise HTTPException(
                400,
                f"Route '{route.route_id}' references unknown stop IDs: {unknown}. "
                "Add them via proposed_stops or upload a new stops.csv.",
            )

    proposed_routes = [r.model_dump() for r in params.proposed_routes]

    current_graph = build_transit_graph(
        _routes, _stops,
        speed_mph=params.average_speed_mph,
        dwell_time=params.dwell_time_minutes,
        transfer_penalty=params.transfer_penalty_minutes,
        highway_speed_mph=params.highway_speed_mph,
        highway_threshold_miles=params.highway_threshold_miles,
    )
    proposed_graph = build_transit_graph(
        proposed_routes, merged_stops,
        speed_mph=params.average_speed_mph,
        dwell_time=params.dwell_time_minutes,
        transfer_penalty=params.transfer_penalty_minutes,
        highway_speed_mph=params.highway_speed_mph,
        highway_threshold_miles=params.highway_threshold_miles,
    )

    return compare_networks(
        current_graph, proposed_graph, merged_stops, _employment_hubs,
        max_travel_minutes=params.max_travel_minutes,
        walking_radius=params.walking_radius_miles,
    )


@protected.get("/api/simulate/coverage-gaps")
def get_coverage_gaps():
    from metrics import identify_coverage_gaps
    gaps = identify_coverage_gaps(_stops)
    return {"gap_count": len(gaps), "gaps": gaps}


@protected.get("/api/route-shapes")
def get_route_shapes():
    """Return GTFS shape coordinates for each route."""
    import json
    shapes_path = os.path.join(DATA_DIR, "route_shapes.json")
    if not os.path.exists(shapes_path):
        return {}
    with open(shapes_path) as f:
        return json.load(f)


# ── Admin-only user management ────────────────────────────────────────────────

class UserCreate(BaseModel):
    username: str
    password: str


@app.get("/api/admin/users", dependencies=[Depends(get_current_admin)])
def admin_list_users():
    return {"users": list_usernames()}


@app.post("/api/admin/users", dependencies=[Depends(get_current_admin)], status_code=201)
def admin_create_user(body: UserCreate):
    if not body.username or not body.password:
        raise HTTPException(status_code=400, detail="Username and password are required")
    if body.username in list_usernames():
        raise HTTPException(status_code=409, detail="Username already exists")
    create_user(body.username, body.password)
    return {"username": body.username}


@app.delete("/api/admin/users/{username}", dependencies=[Depends(get_current_admin)])
def admin_delete_user(username: str):
    if username == "ADMIN":
        raise HTTPException(status_code=400, detail="Cannot delete the ADMIN account")
    if not delete_user(username):
        raise HTTPException(status_code=404, detail="User not found")
    return {"deleted": username}


# ── Register protected router ──────────────────────────────────────────────────
app.include_router(protected)

# ── Serve built React frontend (production / PyInstaller) ─────────────────────
if getattr(_sys, "frozen", False):
    # Bundled exe: static files are in a 'static' folder next to the exe
    STATIC_DIR = os.path.join(os.path.dirname(_sys.executable), "static")
else:
    STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")

if os.path.isdir(STATIC_DIR):
    app.mount("/assets", StaticFiles(directory=os.path.join(STATIC_DIR, "assets")), name="assets")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        return FileResponse(os.path.join(STATIC_DIR, "index.html"))
